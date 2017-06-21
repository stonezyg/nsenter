#!/usr/bin/python

try:
  from openpyxl import Workbook
  from openpyxl.styles import Font
  from openpyxl.styles import colors
except:
  print "OpenPyXL not installed. No EXCEL file output."
  exit(1)

import csv

wb = Workbook()
ws = wb.active
ws.title = 'Baseline Check Results'
row_count = 0

with open('summary.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
	row_count = row_count + 1
	if len(row) == 1:
	    index='A' + str(row_count)
	    ws[index].font = Font(bold=True)
	if len(row) == 3:
	    index='C' + str(row_count)
            if row[2] == 'False':
		ws[index].font = Font(color = colors.RED, bold = True)
            if row[2] == 'N/A':
		ws[index].font = Font(color = 'FF808080', bold = True)
            if row[2] == 'True':
		ws[index].font = Font(color = colors.BLUE, bold = True)

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 10
wb.save('summary.xlsx')


